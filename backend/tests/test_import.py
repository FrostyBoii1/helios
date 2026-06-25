"""Tests for the spreadsheet import staging pipeline (Phase A — parse-only).

Uses a SYNTHETIC in-memory workbook with fabricated names only — no real
customer data, no real workbook, no local paths. Proves: the parser classifies
and flags rows correctly; ingest creates staging rows/issues; and ingest creates
ZERO live Customer/Job records. Also checks admin-only permissions.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.customer import Customer
from app.models.enums import ImportRowClass
from app.models.import_staging import ImportBatch, ImportIssue, ImportRow
from app.models.job import Job
from app.services import import_commit_preview, import_ingest, import_parser

HEADERS = [
    "", "Sales Consultant", "Customer Name", "ADDRESS", "Phone", "Notes",
    "MSB/SB PICS IN FILE?", "Email", "Distributor", "Retailer", "NMI", "Meter No",
    "No of Panels", "Panel Brand/ Wattage", "Inverter Brand/Model", "Storey",
    "Phase", "Roof Type", "Date", "Day", "Time", "Installer",
]


def _synthetic_bytes() -> bytes:
    """A small COMPLETED sheet with fabricated rows covering key parse cases."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPLETED"
    ws.append(HEADERS)  # row 1 = headers

    # row 2 — clean job (known brands -> confident hardware; no Day -> no mismatch)
    ws.append(["TESTIMP0001", "Jane Smith 10/10/2025", "Alex Roe", "1 Test St", "0400000000",
               "", "Yes", "alex@example.test", "Essential", "Origin", "42041234567",
               "M1", "10", "Longi 440", "Goodwe 5kw", "1", "1", "Tin", "", "", "", "Installer One"])
    # row 3 — divider
    ws.append(["FORTNIGHT 1ST-14TH JULY"] + [""] * 21)
    # row 4 — job under divider: multi phone (labelled), multi email, pending approval,
    #         unmatched NMI, uncertain hardware
    ws.append(["TESTIMP0002", "Sales Rep - 30/06/2025", "Pat Lee - PENDING 19/08/2026", "2 Test Rd",
               "0411111111/0422222222 Chris", "", "Yes?", "a@x.test/b@x.test", "Powercor", "AGL",
               "99999999999", "M2", "5", "Brand 415", "mystery unit", "1", "1", "Tile",
               "", "", "", "Installer Two"])
    # row 5 — clean ref but non-name customer cell -> ambiguous_name + approved
    ws.append(["TESTIMP0003", "Jane Smith", "ESSENTIAL APPROVED", "3 Test Ave", "0433333333",
               "", "", "c@x.test", "Essential", "Red Energy", "40011234567", "M3", "8",
               "Brand 415", "Inverter 5kw", "1", "1", "Tin", "", "", "", "Installer One"])
    # row 6 — date/day mismatch (2026-01-01 is Thursday, not Monday)
    ws.append(["TESTIMP0004", "Jane Smith", "Dana Fox", "4 Test Cl", "0444444444", "", "Yes",
               "d@x.test", "Essential", "Origin", "42049999999", "M4", "10", "Brand 440",
               "Inverter 5kw", "1", "1", "Tin", "01/01/2026", "Monday", "08:00", "Installer One"])
    # row 7 — blank
    ws.append([""] * 22)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ws_from_bytes(data: bytes):
    return openpyxl.load_workbook(BytesIO(data), data_only=True)["COMPLETED"]


# --------------------------------------------------------------------------- #
# Parser unit tests (pure, DB-free)
# --------------------------------------------------------------------------- #
def test_parser_classifies_and_parses():
    rows = list(import_parser.parse_rows(_ws_from_bytes(_synthetic_bytes())))
    by_class: dict[str, list] = {}
    for r in rows:
        by_class.setdefault(r.row_class, []).append(r)

    assert len(by_class["job"]) == 4
    assert len(by_class["divider"]) == 1
    assert len(by_class["blank"]) == 1

    clean = next(r for r in rows if r.legacy_reference == "TESTIMP0001")
    assert clean.parsed["customer_name"] == "Alex Roe"
    assert clean.parsed["address"] == "1 Test St"  # address now in parsed candidate
    assert clean.parsed["salesperson"] == "Jane Smith"
    assert clean.parsed["sale_date"] == "10/10/2025"
    assert clean.parsed["distributor_inferred"] == "NSW Essential"  # NMI 4204
    assert clean.parsed["msb_state"] == "yes"
    assert clean.issues == []


# --------------------------------------------------------------------------- #
# R2: blank / near-blank short-circuit. A row whose mapped fields are ALL empty
# is BLANK (parsed {}, no issues) even when stray/unmapped noise cells exist;
# sparse rows with meaningful mapped data are never swallowed as blank.
# --------------------------------------------------------------------------- #
EMPTY_ROW = [""] * len(HEADERS)


def _parse_data_rows(data_rows: list[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPLETED"
    ws.append(HEADERS)
    for row in data_rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return list(import_parser.parse_rows(_ws_from_bytes(buf.getvalue())))


def test_truly_blank_row_parses_empty():
    rows = _parse_data_rows([EMPTY_ROW])
    assert len(rows) == 1
    b = rows[0]
    assert b.row_class == "blank"
    assert b.parsed == {}
    assert b.issues == []


def test_near_blank_row_with_stray_noise_is_blank():
    # Every mapped field empty; a stray value in a HEADER-LESS far column is pure noise that used to
    # inflate the cell count -> 'ambiguous' + a spurious ambiguous_name error. Now it stays blank.
    rows = _parse_data_rows([EMPTY_ROW + ["stray noise cell"]])
    assert len(rows) == 1
    b = rows[0]
    assert b.row_class == "blank"
    assert b.parsed == {}
    assert b.issues == []
    assert all(i["kind"] != "ambiguous_name" for i in b.issues)


def test_near_blank_row_with_unmapped_column_is_blank():
    # A value in an extra UNMAPPED (but headered) column is captured in raw["_unmapped"] for
    # traceability but must NOT promote the row to a job.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPLETED"
    ws.append(HEADERS + ["Mystery Extra Column"])
    ws.append(EMPTY_ROW + ["leftover value"])
    buf = BytesIO()
    wb.save(buf)
    rows = list(import_parser.parse_rows(_ws_from_bytes(buf.getvalue())))
    assert len(rows) == 1
    assert rows[0].row_class == "blank"
    assert rows[0].parsed == {}
    assert rows[0].issues == []
    assert rows[0].raw.get("_unmapped") == {"Mystery Extra Column": "leftover value"}


def test_sparse_real_row_with_only_name_not_blank():
    only_name = list(EMPTY_ROW)
    only_name[2] = "Jordan Real"  # Customer Name (mapped) -> meaningful data
    rows = _parse_data_rows([only_name])
    assert len(rows) == 1
    assert rows[0].row_class != "blank"
    assert rows[0].parsed.get("customer_name") == "Jordan Real"


def test_sparse_real_row_with_only_nmi_not_blank():
    only_nmi = list(EMPTY_ROW)
    only_nmi[10] = "42041234567"  # NMI (mapped) -> meaningful data
    rows = _parse_data_rows([only_nmi])
    assert len(rows) == 1
    assert rows[0].row_class != "blank"


def test_ingest_near_blank_row_stored_as_blank(db_session: Session, users):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPLETED"
    ws.append(HEADERS)
    job = list(EMPTY_ROW)
    job[0], job[2], job[3], job[10] = "SC1234", "Real Person", "1 Real St", "42041234567"
    ws.append(job)                          # a real job
    ws.append(EMPTY_ROW + ["stray noise"])  # near-blank: stray value in a far column
    buf = BytesIO()
    wb.save(buf)
    batch = import_ingest.ingest_bytes(
        db_session,
        file_bytes=buf.getvalue(),
        source_filename="nearblank.xlsx",
        created_by_id=users["admin"].id,
    )
    db_session.flush()
    assert batch.job_rows == 1
    assert batch.blank_rows == 1
    assert batch.ambiguous_rows == 0  # the near-blank row is NOT classified as an ambiguous job
    rows = db_session.scalars(
        select(ImportRow).where(ImportRow.batch_id == batch.id).order_by(ImportRow.source_row_index)
    ).all()
    blank = next(r for r in rows if r.row_class == ImportRowClass.BLANK.value)
    assert blank.parsed is None  # parse-level {} is stored as None (no parsed junk to review)
    assert blank.raw is not None  # raw cells preserved for traceability
    n_issues = db_session.scalar(
        select(func.count()).select_from(ImportIssue).where(ImportIssue.row_id == blank.id)
    )
    assert n_issues == 0
    # Commit preview excludes blanks as blank_or_divider.
    assert import_commit_preview.classify_row(blank) == "blank_or_divider"


def test_parser_flags_issues():
    rows = list(import_parser.parse_rows(_ws_from_bytes(_synthetic_bytes())))

    r2 = next(r for r in rows if r.legacy_reference == "TESTIMP0002")
    kinds2 = {i["kind"] for i in r2.issues}
    assert {"multi_phone", "multi_email", "nmi_unmatched"} <= kinds2
    assert r2.parsed["customer_name"] == "Pat Lee"
    assert r2.parsed["approval_state"] == "pending"  # "PENDING 19/08/2026"
    assert r2.parsed["approval_pending_date"] == "19/08/2026"
    assert r2.context_text == "FORTNIGHT 1ST-14TH JULY"  # carried from the divider
    # phone label kept only because it was explicit
    assert any(p["label"] == "Chris" for p in r2.parsed["phones"])

    r3 = next(r for r in rows if r.legacy_reference == "TESTIMP0003")
    assert any(i["kind"] == "ambiguous_name" for i in r3.issues)
    assert r3.parsed["approval_state"] == "approved"

    r4 = next(r for r in rows if r.legacy_reference == "TESTIMP0004")
    assert any(i["kind"] == "date_day_mismatch" for i in r4.issues)


def test_parse_rows_derives_needs_approval_from_panels_inverter():
    """R2: parse_rows derives approval_state='required' (Needs approval) for a
    numeric-panel + inverter job with no explicit approval evidence — matching the
    commit-time auto-label — without overriding approved/pending/action-phrase, and
    without firing for battery-only / no-panel / inverter-only / non-numeric / zero."""
    def _row(ref, name, panels, inverter):
        cells = [""] * len(HEADERS)
        cells[0], cells[2], cells[3] = ref, name, "1 Test St"
        cells[10] = "42041234567"           # NMI (healthy job, distributor matches)
        cells[12], cells[14] = panels, inverter  # No of Panels, Inverter
        return cells

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPLETED"
    ws.append(HEADERS)
    ws.append(_row("SC9001", "Solar Sam", "12", "Goodwe 5kw"))                  # -> required (derived)
    ws.append(_row("SC9002", "Amy Jones - ESSENTIAL APPROVED", "10", "Goodwe 5kw"))  # approved kept
    ws.append(_row("SC9003", "Pete Smith - PENDING 19/08/2026", "10", "Goodwe 5kw"))  # pending kept
    ws.append(_row("SC9004", "Andy Brown - DO APPROVAL", "10", "Goodwe 5kw"))   # required (action phrase)
    ws.append(_row("SC9005", "Bob Battery", "", "Goodwe 5kw"))                  # inverter-only -> none
    ws.append(_row("SC9006", "Ned NoPanel", "-", "Goodwe 5kw"))                 # non-numeric panels -> none
    ws.append(_row("SC9007", "Zoe Zero", "0", "Goodwe 5kw"))                    # zero panels -> none
    ws.append(_row("SC9008", "Pam PanelsOnly", "10", ""))                      # no inverter -> none
    buf = BytesIO()
    wb.save(buf)
    rows = {r.legacy_reference: r for r in import_parser.parse_rows(_ws_from_bytes(buf.getvalue()))}

    assert rows["SC9001"].parsed["approval_state"] == "required"   # numeric panels + inverter
    assert rows["SC9002"].parsed["approval_state"] == "approved"   # explicit approved not overridden
    assert rows["SC9003"].parsed["approval_state"] == "pending"    # explicit pending not overridden
    assert rows["SC9004"].parsed["approval_state"] == "required"   # explicit action phrase kept
    for ref in ("SC9005", "SC9006", "SC9007", "SC9008"):
        assert rows[ref].parsed["approval_state"] == "none", ref

    # The derived review state and the commit-time label agree for the same row
    # (shared needs_approval_from_panels predicate — they cannot diverge).
    from app.services.job_labels import auto_label_keys
    p = rows["SC9001"].parsed
    assert ("approval_required", None) in auto_label_keys(p, p.get("details"))


# --------------------------------------------------------------------------- #
# Name-cell trailing notes + decommission detection (pure helpers)
# --------------------------------------------------------------------------- #
def test_clean_name_cell_notes_strips_approval_keeps_meaning():
    # Meaningful operational text is preserved verbatim.
    assert import_parser.clean_name_cell_notes("includes hot water timer") == "includes hot water timer"
    assert (
        import_parser.clean_name_cell_notes("undersold Brighte fees, check after install")
        == "undersold Brighte fees, check after install"
    )
    # Pure approval status leaves nothing meaningful behind.
    assert import_parser.clean_name_cell_notes("APPROVED") == ""
    assert import_parser.clean_name_cell_notes("PENDING 19/08/2026") == ""
    # Approval mixed with a real note keeps only the real note.
    assert (
        import_parser.clean_name_cell_notes("APPROVED includes hot water timer")
        == "includes hot water timer"
    )
    assert import_parser.clean_name_cell_notes("") == ""


def test_clean_name_cell_notes_strips_network_approval_residue():
    cn = import_parser.clean_name_cell_notes

    # The three documented examples: approval phrase (incl. network label) gone,
    # meaningful notes before/after preserved, no bare network residue.
    r1 = cn("includes hot water timer - ESSENTIAL APPROVED - Report Ref # - 96004-Y")
    assert "ESSENTIAL" not in r1.upper()
    assert "includes hot water timer" in r1 and "Report Ref" in r1 and "96004-Y" in r1

    r2 = cn("undersold Brighte fees, check after install - ESSENTIAL APPROVED")
    assert "ESSENTIAL" not in r2.upper()
    assert r2 == "undersold Brighte fees, check after install"

    r3 = cn("ESSENTIAL APPROVED - prescreened wk24/2 - REMOVE OLD SYSTEM")
    assert "ESSENTIAL" not in r3.upper()
    # Phase-7 fix: the decommission marker is stripped from name-cell notes (the
    # remove-old-system flag is detected separately); meaningful note text remains.
    assert "prescreened wk24/2" in r3 and "REMOVE OLD SYSTEM" not in r3.upper()

    # Other network labels are stripped as part of an APPROVED phrase.
    for net in ("ENERGEX", "ERGON", "ENDEAVOUR", "AUSGRID", "AUSNET",
                "POWERCOR", "UNITED", "JEMENA", "SAPN"):
        out = cn(f"keep this - {net} APPROVED")
        assert net not in out.upper(), net
        assert out == "keep this"

    # PENDING-with-date phrases (incl. a network label) are removed; note kept.
    assert cn("JEMENA PENDING 19/08/2026 - call customer") == "call customer"
    assert cn("ERGON PENDING 01/02/2025") == ""

    # Case-insensitive.
    assert "ESSENTIAL" not in cn("note - essential approved").upper()

    # Phase-7 fix: decommission markers are stripped from name-cell notes (the
    # flag is set separately); an approval+marker-only cell leaves nothing.
    assert cn("ENDEAVOUR APPROVED - REMOVE OLD SYSTEM") == ""

    # A network word that is NOT part of an approval status is meaningful content
    # and must be kept.
    assert cn("ESSENTIAL repairs needed") == "ESSENTIAL repairs needed"
    assert cn("needs ENERGEX meter upgrade") == "needs ENERGEX meter upgrade"


def test_clean_name_cell_notes_strips_decommission_marker_keeps_dates_and_text():
    cn = import_parser.clean_name_cell_notes
    # Marker-only -> nothing left (the flag is detected separately).
    assert cn("REMOVE OLD SYSTEM") == ""
    assert cn("DECOM") == ""
    # Marker stripped; a standalone date is PRESERVED verbatim (never inferred).
    out = cn("remove old system - 15/04/1983")
    assert "REMOVE OLD SYSTEM" not in out.upper()
    assert "15/04/1983" in out
    # Marker stripped; meaningful note preserved.
    out2 = cn("includes hot water timer - DECOM")
    assert "DECOM" not in out2.upper()
    assert "includes hot water timer" in out2
    # The remove-old-system flag is still detected independently from the raw text.
    assert import_parser.detect_decommission("remove old system - 15/04/1983") is not None


def test_detect_decommission_variants():
    for text in (
        "REMOVE OLD SYSTEM",
        "remove old system",
        "DECOM",
        "decommission",
        "decommision",   # missing one 's'
        "decomission",   # missing one 'm'
        "needs DECOM before install",
        "Customer wants to remove old system first",
    ):
        assert import_parser.detect_decommission(text) is not None, text
    # No false positives on ordinary text.
    assert import_parser.detect_decommission("Alex Roe", "includes hot water timer") is None
    assert import_parser.detect_decommission("") is None
    # Returns the matched marker text for reviewer visibility.
    assert import_parser.detect_decommission("Jane DECOM").upper() == "DECOM"


def _one_job_row_bytes(
    *, name_cell: str, notes_cell: str = "", sales_cell: str = "Rep 10/10/2025",
    panels: str = "10",
) -> bytes:
    """A minimal COMPLETED sheet: header + one clean job row, with the Customer
    Name + Notes + Sales Consultant + No-of-Panels cells controllable. Fabricated
    data only."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPLETED"
    ws.append(HEADERS)
    ws.append(["TESTIMP0009", sales_cell, name_cell, "9 Test St", "0400000000",
               notes_cell, "Yes", "x@example.test", "Essential", "Origin", "42041234567",
               "M9", panels, "Longi 440", "Goodwe 5kw", "1", "1", "Tin", "", "", "", "Installer One"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_rows_preserves_name_cell_notes():
    rows = list(import_parser.parse_rows(_ws_from_bytes(
        _one_job_row_bytes(name_cell="Pat Lee - includes hot water timer")
    )))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    assert row.parsed["customer_name"] == "Pat Lee"
    assert row.parsed["customer_name_notes"] == "includes hot water timer"
    assert row.parsed["removes_old_system"] is False
    assert row.parsed["decommission_marker"] is None


def test_parse_rows_flags_decommission_from_name_or_notes():
    # Marker in the name cell (no stop-marker preceding it).
    rows = list(import_parser.parse_rows(_ws_from_bytes(
        _one_job_row_bytes(name_cell="Sam Roe DECOM")
    )))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    assert row.parsed["removes_old_system"] is True
    assert row.parsed["decommission_marker"].upper() == "DECOM"

    # Marker in the Notes column.
    rows2 = list(import_parser.parse_rows(_ws_from_bytes(
        _one_job_row_bytes(name_cell="Dana Fox", notes_cell="REMOVE OLD SYSTEM before install")
    )))
    row2 = next(r for r in rows2 if r.legacy_reference == "TESTIMP0009")
    assert row2.parsed["removes_old_system"] is True
    assert "remove old system" in row2.parsed["decommission_marker"].lower()


def test_parse_customer_name_strips_decommission_marker():
    cn = import_parser.parse_customer_name
    # Marker glued to the name (no " - " stop) is stripped from the name.
    assert cn("Jane Roe -remove old system")["name"] == "Jane Roe"
    assert cn("Sam Roe DECOM")["name"] == "Sam Roe"
    # Marker after a " - " stop: the name is already clean.
    assert cn("Customer Name - remove old system")["name"] == "Customer Name"
    # Marker-only cell -> empty name, not treated as a name.
    only = cn("remove old system")
    assert only["name"] == "" and only["looks_like_name"] is False
    # Non-decommission names are unchanged.
    assert cn("Pat Lee")["name"] == "Pat Lee"


def test_parse_rows_strips_decommission_marker_from_name_keeps_flag():
    # End-to-end: marker glued to the name must NOT remain in customer_name, while
    # the remove-old-system flag + marker are still detected.
    rows = list(import_parser.parse_rows(_ws_from_bytes(
        _one_job_row_bytes(name_cell="Jane Roe -remove old system")
    )))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    assert row.parsed["customer_name"] == "Jane Roe"
    assert row.parsed["removes_old_system"] is True
    assert "remove old system" in row.parsed["decommission_marker"].lower()
    assert "remove old system" not in (row.parsed.get("customer_name_notes") or "").lower()


# --------------------------------------------------------------------------- #
# Phase-7 parser fixes (batch 3299 owner review): Sales Consultant suffix split,
# Customer Name land/legal descriptor split, name-cell distributor approval phrase.
# --------------------------------------------------------------------------- #
def test_parse_sales_consultant_splits_non_name_suffix():
    psc = import_parser.parse_sales_consultant
    # payment suffix -> name kept, suffix preserved, no sale date
    r = psc("Jason Gowans - cash")
    assert r["name"] == "Jason Gowans"
    assert r["misfiled"] == "cash"
    assert r["sale_date"] is None
    # system + payment suffix
    r2 = psc("Jason G - 13.28kw Humm")
    assert r2["name"] == "Jason G"
    assert r2["misfiled"] == "13.28kw Humm"
    # a labelled DOB suffix is preserved verbatim, NEVER a sale_date, no DOB field
    r3 = psc("Robert W - dob 14/05/1980")
    assert r3["name"] == "Robert W"
    assert r3["misfiled"] == "dob 14/05/1980"
    assert r3["sale_date"] is None


def test_parse_sales_consultant_keeps_clean_names_and_bare_dates():
    psc = import_parser.parse_sales_consultant
    # bare-date suffix is still the sale date (prior behaviour preserved)
    r = psc("Jane Smith - 10/10/2025")
    assert r["name"] == "Jane Smith" and r["sale_date"] == "10/10/2025" and r["misfiled"] is None
    # trailing date without separator is unchanged
    r2 = psc("Jane Smith 10/10/2025")
    assert r2["name"] == "Jane Smith" and r2["sale_date"] == "10/10/2025" and r2["misfiled"] is None
    # a plain name is untouched
    r3 = psc("Jane Smith")
    assert r3 == {"name": "Jane Smith", "sale_date": None, "misfiled": None}


def test_parse_sales_consultant_extracts_leading_suffix_date():
    # A MIXED suffix that LEADS with a plain date: the date becomes the sale date,
    # the remainder (a labelled DOB note) is preserved verbatim. No DOB is inferred.
    psc = import_parser.parse_sales_consultant
    r = psc("Robert W - 4/4/2023 - dob 23/11/55")
    assert r["name"] == "Robert W"
    assert r["sale_date"] == "4/4/2023"
    assert r["misfiled"] == "dob 23/11/55"
    # downstream this resolves to the expected ISO sale date.
    assert import_parser.parse_date_maybe(r["sale_date"]).isoformat() == "2023-04-04"
    # A labelled date with NO leading plain date stays whole, never a sale date.
    r2 = psc("Robert W - dob 23/11/55")
    assert r2["sale_date"] is None and r2["misfiled"] == "dob 23/11/55"


def test_parse_customer_name_splits_land_descriptor():
    cn = import_parser.parse_customer_name
    # hyphen-glued Lot/DP (the src-88 shape) -> name clean, descriptor preserved
    r = cn("Jane Roe -Lot 4 DP 588479")
    assert r["name"] == "Jane Roe"
    assert r["land_descriptor"] == "Lot 4 DP 588479"
    # space-separated Lot/DP
    r2 = cn("John Smith - Lot 7 DP 12345")
    assert r2["name"] == "John Smith"
    assert r2["land_descriptor"] == "Lot 7 DP 12345"
    # no false positive: a surname containing "lot" and no parcel number
    r3 = cn("Charlotte Lott")
    assert r3["name"] == "Charlotte Lott"
    assert r3["land_descriptor"] is None


def test_parse_customer_name_separates_distributor_approval_phrase():
    cn = import_parser.parse_customer_name
    # reference phrase (the src-61 shape) -> phrase out of the name, preserved
    r = cn("Jane Roe - Jemena Approval # 000413493")
    assert r["name"] == "Jane Roe"
    assert "jemena" not in r["name"].lower() and "approval" not in r["name"].lower()
    assert r["approval_phrase"] == "Jemena Approval # 000413493"
    # glued by hyphen-space (the src-323 shape) -> network label not left in name
    r2 = cn("John Smith- JEMENA APPROVAL #000445604")
    assert r2["name"] == "John Smith"
    assert "jemena" not in r2["name"].lower()
    # a status word in the phrase still drives approval_state via parse_approval
    appr = import_parser.parse_approval(cn("Pat Lee ERGON APPROVED")["approval_phrase"] or "")
    assert appr["state"] == "approved"
    # a network word that is NOT an approval phrase is left in place, not stripped
    r3 = cn("Bob Lee - ENERGEX meter upgrade needed")
    assert r3["approval_phrase"] is None
    assert r3["name"] == "Bob Lee"


def test_parse_approval_reference_means_approved():
    pa = import_parser.parse_approval
    # A distributor approval REFERENCE number means the connection was approved.
    assert pa("Jemena Approval # 000413493")["state"] == "approved"
    assert pa("ENERGEX APPROVAL No 12345")["state"] == "approved"
    assert pa("Approval Ref 678901")["state"] == "approved"
    # An explicit APPROVED word still wins.
    assert pa("ESSENTIAL APPROVED")["state"] == "approved"
    # PENDING wins over a trailing reference -> pending (not approved), date kept.
    p = pa("ENERGEX PENDING 19/08/2026 ref 123456")
    assert p["state"] == "pending" and p["pending_date"] == "19/08/2026"
    # No evidence -> none (no approval label). Digit runs that are NOT approval
    # references (a DOB date, a 4-digit year with no #/No/Ref marker) do not flip.
    assert pa("")["state"] == "none"
    assert pa("Robert W dob 23/11/1955")["state"] == "none"
    assert pa("install expected 2025")["state"] == "none"
    assert pa("approval expected 2025")["state"] == "none"


def test_parse_approval_action_phrase_means_required():
    # R2/A3: an explicit instruction to OBTAIN approval ("DO APPROVAL", "NEEDS
    # APPROVAL", ...) classifies as approval_required ("Needs approval"), NOT
    # approved. The owner's full keyword set is covered.
    pa = import_parser.parse_approval
    for txt in ("DO APPROVAL", "NEED APPROVAL", "NEEDS APPROVAL", "APPLY APPROVAL",
                "ORGANISE APPROVAL", "ORGANIZE APPROVAL", "GET APPROVAL",
                "approval needed", "approval required"):
        assert pa(txt)["state"] == "required", txt
    # An action phrase wins over a stray APPROVED word elsewhere (still not done).
    assert pa("NEEDS APPROVAL")["state"] == "required"
    # A past-tense "APPROVED" with no action verb stays approved.
    assert pa("ESSENTIAL APPROVED")["state"] == "approved"
    # A real reference number still means approved (no action verb).
    assert pa("Jemena Approval # 000413493")["state"] == "approved"


def test_clean_name_cell_notes_strips_action_phrase_keeps_context():
    # R2/A3: the bare action phrase is dropped (its meaning is on the label); any
    # surrounding operational context is preserved verbatim.
    cn = import_parser.clean_name_cell_notes
    assert cn("DO APPROVAL") == ""
    assert cn("NEEDS APPROVAL JEMENA") == "JEMENA"
    assert "TECHNAUS POWERCOR PORTAL" in cn("DO APPROVAL TECHNAUS POWERCOR PORTAL - BACKSTOP")
    assert "do approval" not in cn("Jason Check Solax and phase and do approval").lower()
    assert "Jason Check Solax and phase" in cn("Jason Check Solax and phase and do approval")


# --------------------------------------------------------------------------- #
# P3 parser fixes (batch 5846 owner review): non-name suffixes peeled off the
# Customer Name cell + an email-only name is a blocking error. Each is a CATEGORY,
# not a one-off: the suffix is preserved VERBATIM as a name-cell note (no text
# loss, no structured DOB) and the name is left clean.
# --------------------------------------------------------------------------- #
def test_parse_customer_name_strips_dob_phrase():
    cn = import_parser.parse_customer_name
    # "DATE OF BIRTH" hyphen-glued (no space) -> name cleaned, text preserved.
    r = cn("Cooper Boardman -DATE OF BIRTH 10/04/2003")
    assert r["name"] == "Cooper Boardman"
    assert r["extracted"] == "DATE OF BIRTH 10/04/2003"
    # lowercase "dob" + a date.
    r2 = cn("Margaret Sutton -dob 15/11/1955")
    assert r2["name"] == "Margaret Sutton"
    assert r2["extracted"] == "dob 15/11/1955"
    # NO structured DOB field is ever produced.
    assert "dob" not in r and "date_of_birth" not in r
    # A "Dobson" surname is NOT stripped (needs the DOB label followed by a date).
    assert cn("John Dobson")["name"] == "John Dobson"


def test_parse_customer_name_strips_bare_trailing_date():
    cn = import_parser.parse_customer_name
    # hyphen-glued bare date.
    r = cn("Naomi Carter- 18/4/75")
    assert r["name"] == "Naomi Carter"
    assert r["extracted"] == "18/4/75"
    # space-only trailing date after a two-person name -> BOTH names kept.
    r2 = cn("Edward Joshua- Claire Joshua 11/05/2003")
    assert r2["name"] == "Edward Joshua- Claire Joshua"
    assert "11/05/2003" in r2["extracted"]
    # no false strip: a plain name is untouched.
    assert cn("Pat Lee")["name"] == "Pat Lee"


def test_parse_customer_name_strips_midcell_bare_date():
    # R2/A1: a bare date glued to the name with MORE text after it (a licence / lot /
    # state tail) is split off — the end-anchored bare-date rule misses these because
    # the date no longer ends the cell. The date is PRESERVED, never inferred as DOB.
    cn = import_parser.parse_customer_name
    # The src-203 shape: "Naomi Carter- 18/4/75 - DL - 11878134 - Exp ... - NSW".
    r = cn("Naomi Carter- 18/4/75 - DL - 11878134 - Exp 02/03/2028 - NSW")
    assert r["name"] == "Naomi Carter"
    assert "18/4/75" in r["extracted"]
    assert "DL - 11878134" in r["extracted"]
    # The src-308 shape: two-person name then a DOB then more text.
    r2 = cn("Richard and Teresa Simmons- 28/02/1948 Richard - 16/07/1952 - teresa - NETWORK APPROVED")
    assert r2["name"] == "Richard and Teresa Simmons"
    assert "28/02/1948" in r2["extracted"]
    # No structured DOB field is ever produced.
    assert "dob" not in r and "date_of_birth" not in r
    # Guard intact: a PENDING approval date stays adjacent to its keyword (not cut as
    # a bare note), so parse_approval can still read the pending date.
    r3 = cn("Pat Lee - ENERGEX PENDING 19/08/2026")
    assert r3["name"] == "Pat Lee"
    appr = import_parser.parse_approval(r3.get("approval_phrase") or "", r3["extracted"])
    assert appr["state"] == "pending" and appr["pending_date"] == "19/08/2026"


# --------------------------------------------------------------------------- #
# R2 name-suffix cleanup — operational/source-context suffixes are pattern
# FAMILIES, not name-specific hacks. The name is cleaned; the exact stripped text
# is preserved (it flows to customer_name_notes -> On Commit / Job internal notes).
# A space-BEFORE-hyphen (" -note") is an unambiguous note delimiter; a hyphen-THEN-
# space ("- note") only strips when a recognized family keyword follows, so entity
# appositives ("Inn- Wayne Bond") survive.
# --------------------------------------------------------------------------- #
def _name_note(raw):
    r = import_parser.parse_customer_name(raw)
    return r["name"], import_parser.clean_name_cell_notes(r["extracted"])


def test_suffix_family_booked_prescreened():
    # Category 1: booked / prescreened scheduling suffixes (date, weekday, range, wk).
    for raw, name, frag in [
        ("Paul Neilsen and Carly Sorenson -booked 28/8", "Paul Neilsen and Carly Sorenson", "booked 28/8"),
        ("Kylie Felton- booked 4/10", "Kylie Felton", "booked 4/10"),
        ("Mal Moody- booked 06/12", "Mal Moody", "booked 06/12"),
        ("Gideon Vos -booked 5 or 6/2", "Gideon Vos", "booked"),
        ("Lee Bargwanna -prescreened wk 1/9", "Lee Bargwanna", "prescreened"),
        ("Janelle and Wayne Alexander- prescreened wk 17/11", "Janelle and Wayne Alexander", "prescreened"),
    ]:
        n, note = _name_note(raw)
        assert n == name, raw
        assert frag in note, (raw, note)


def test_suffix_family_admin_social_source():
    # Category 2: vm, on fb, POLE, AGREED, SV submitted.
    for raw, name, frag in [
        ("Wendy and Matthew Loffler- vm", "Wendy and Matthew Loffler", "vm"),
        ("Chantal Jackson -on fb", "Chantal Jackson", "on fb"),
        ("Heather and George Case -POLE", "Heather and George Case", "POLE"),
        ("Allen Irwen- AGREED", "Allen Irwen", "AGREED"),
        ("Kerryn Robertson -SV submitted 17/5", "Kerryn Robertson", "SV submitted 17/5"),
    ]:
        n, note = _name_note(raw)
        assert n == name, raw
        assert frag in note, (raw, note)


def test_suffix_family_export_system():
    # Category 3: export / Nkw export system suffix.
    n, note = _name_note("Kathleen Jones -8kw export")
    assert n == "Kathleen Jones"
    assert "8kw export" in note


def test_suffix_family_multipart_freeform_note():
    # Category 4: a free-form job/admin note after a space-before-hyphen is kept whole
    # (including a second internal hyphen), never part of the name.
    n, note = _name_note("Peter and Lesley Wenselowski -Jason check wiring to shed- hot water timer")
    assert n == "Peter and Lesley Wenselowski"
    assert note == "Jason check wiring to shed- hot water timer"
    # A single-name shorthand after a couple name is context, not part of the name.
    n2, note2 = _name_note("Kelly Double and Troy McGillivray -Troy")
    assert n2 == "Kelly Double and Troy McGillivray"
    assert note2 == "Troy"


def test_suffix_family_invoice_sent_keeps_entity_contact():
    # Category 5: an invoice note on a business/hotel entity — only the note is
    # stripped; the entity + its contact appositive ("Inn- Wayne Bond") survive.
    n, note = _name_note("The Leeton Heritage Motor Inn- Wayne Bond- 2 invoices sent")
    assert n == "The Leeton Heritage Motor Inn- Wayne Bond"
    assert "2 invoices sent" in note


def test_suffix_family_trailing_empty_delimiter():
    # Category 6: a trailing bare " -" delimiter is dropped; no note, clean name.
    for raw, name in [("Sarah and Barry Pitkin -", "Sarah and Barry Pitkin"), ("Wayne Giles -", "Wayne Giles")]:
        n, note = _name_note(raw)
        assert n == name and note == "", raw


def test_suffix_family_conservative_on_entities_and_names():
    # Category 7 + guards: company/trust names and hyphenated surnames are NOT
    # rewritten unless a confident delimiter+keyword pattern exists.
    cn = import_parser.parse_customer_name
    # Trust/company entity with no hyphen-with-space and no family keyword -> untouched
    # (manual-resolution case; the workbook has no source field naming "Jules").
    horton = "C &J Horton PTY as Trustees for Horton Family Superrannuation Fund"
    assert cn(horton)["name"] == horton
    # A hyphenated surname (no spaces around the hyphen) is never split.
    assert cn("Smith-Pole")["name"] == "Smith-Pole"
    assert cn("Brenton Hoskins-Murphy -booked 4/11- AGREED WITH UPGRADE")["name"] == "Brenton Hoskins-Murphy"
    # A real surname token after a plain space (no hyphen) is never stripped.
    assert cn("Mary Pole")["name"] == "Mary Pole"
    # A company name with a place suffix after a space-before-hyphen IS cleaned.
    assert cn("Grow Nuts Pty Ltd -Griffith")["name"] == "Grow Nuts Pty Ltd"


def test_parse_customer_name_strips_pillar_reference():
    cn = import_parser.parse_customer_name
    r = cn("Steve Olive pillar 111178023")
    assert r["name"] == "Steve Olive"
    assert r["extracted"] == "pillar 111178023"
    # a surname "Pillar" with no following number is never stripped.
    assert cn("Joan Pillar")["name"] == "Joan Pillar"


def test_parse_customer_name_strips_export_limited():
    cn = import_parser.parse_customer_name
    r = cn("Tracy Bain 2.28KW EXPORT LIMITED")
    assert r["name"] == "Tracy Bain"
    assert r["extracted"] == "2.28KW EXPORT LIMITED"
    # "kw" inside a surname is not an export-limit annotation.
    assert cn("Mark Kwan")["name"] == "Mark Kwan"


def test_parse_customer_name_strips_finalise_to():
    cn = import_parser.parse_customer_name
    r = cn("Michael Simpson FINALISE TO AGL")
    assert r["name"] == "Michael Simpson"
    assert r["extracted"] == "FINALISE TO AGL"
    # the whole trailing instruction (incl. a later date) is preserved as one note.
    r2 = cn("Michael Simpson  FINALISE TO AGL - GET HIS BILL DETAILS - 06/02/64")
    assert r2["name"] == "Michael Simpson"
    assert r2["extracted"] == "FINALISE TO AGL - GET HIS BILL DETAILS - 06/02/64"
    # P3 does NOT auto-label admin work — approval status stays 'none' (that is P4).
    assert import_parser.parse_approval(r["extracted"])["state"] == "none"


def test_parse_customer_name_email_only_flagged():
    cn = import_parser.parse_customer_name
    r = cn("jjmckoz82@gmail.com")
    assert r["email_only"] is True
    assert r["looks_like_name"] is False
    # a mixed "Name <email>" cell is NOT email-only (left untouched in this pass).
    assert cn("Jane Smith jjmckoz82@gmail.com")["email_only"] is False
    # an empty cell is not email-only.
    assert cn("")["email_only"] is False


def test_parse_rows_email_only_name_is_blocking_error():
    rows = list(import_parser.parse_rows(_ws_from_bytes(
        _one_job_row_bytes(name_cell="jjmckoz82@gmail.com")
    )))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    kinds = {(i["kind"], i["severity"]) for i in row.issues}
    assert ("email_only_name", "error") in kinds          # blocking (error severity)
    assert not any(i["kind"] == "ambiguous_name" for i in row.issues)  # specific kind, not generic


def test_parse_rows_name_suffix_lands_in_customer_name_notes():
    # End-to-end: each non-name suffix is peeled from customer_name and preserved in
    # customer_name_notes -> details.notes.customer_name_notes, the bucket that seeds
    # Job.internal_notes after commit (P2). No structured DOB; no text loss.
    cases = {
        "Cooper Boardman -DATE OF BIRTH 10/04/2003": ("Cooper Boardman", "DATE OF BIRTH 10/04/2003"),
        "Steve Olive pillar 111178023": ("Steve Olive", "pillar 111178023"),
        "Tracy Bain 2.28KW EXPORT LIMITED": ("Tracy Bain", "2.28KW EXPORT LIMITED"),
        "Michael Simpson FINALISE TO AGL": ("Michael Simpson", "FINALISE TO AGL"),
        "Naomi Carter- 18/4/75": ("Naomi Carter", "18/4/75"),
        "Margaret Sutton -dob 15/11/1955": ("Margaret Sutton", "dob 15/11/1955"),
    }
    for name_cell, (exp_name, exp_note) in cases.items():
        rows = list(import_parser.parse_rows(_ws_from_bytes(
            _one_job_row_bytes(name_cell=name_cell)
        )))
        row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
        assert row.parsed["customer_name"] == exp_name, name_cell
        assert row.parsed["customer_name_notes"] == exp_note, name_cell
        # flows into the structured notes bucket that internal_notes seeding reads.
        assert row.parsed["details"]["notes"]["customer_name_notes"] == exp_note, name_cell
        # no structured DOB / approval invented from a date-bearing suffix.
        assert "approval" not in row.parsed["details"], name_cell


def test_parse_rows_jemena_reference_is_approved_and_preserved():
    # End-to-end: a name-cell "Jemena Approval # …" -> approval_state approved AND
    # the exact reference is kept as a neutral review note (never lost).
    rows = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Jane Roe - Jemena Approval # 000413493",
    ))))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    assert row.parsed["approval_state"] == "approved"
    assert row.parsed["customer_name"] == "Jane Roe"
    review = [
        m["text"] for m in row.parsed["details"]["notes"].get("review_notes", [])
        if m["source_column"] == "Customer Name"
    ]
    assert any("000413493" in t for t in review)


def test_parse_rows_routes_suffixes_to_correct_notes_buckets():
    # End-to-end through parse_rows -> build_details. A land/legal parcel descriptor
    # stays a (neutral) misfiled SOURCE note; a sales-cell free-note and a
    # distributor approval phrase are neutral REVIEW notes. Structured name /
    # salesperson stay clean.
    rows = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Jane Roe -Lot 4 DP 588479",
        sales_cell="Jason Gowans - cash",
    ))))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    assert row.parsed["customer_name"] == "Jane Roe"
    assert row.parsed["salesperson"] == "Jason Gowans"
    notes = row.parsed["details"]["notes"]
    misfiled = {(m["source_column"], m["text"]) for m in notes.get("misfiled", [])}
    review = {(m["source_column"], m["text"]) for m in notes.get("review_notes", [])}
    # Land descriptor -> misfiled "Customer Name"; sales free-note -> review_notes.
    assert ("Customer Name", "Lot 4 DP 588479") in misfiled
    assert ("Sales Consultant", "cash") in review
    assert ("Sales Consultant", "cash") not in misfiled

    # Name-cell distributor approval REFERENCE phrase (src-61 shape): out of the
    # name, preserved as a neutral REVIEW note, NOT a misfiled warning.
    rows2 = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Jane Roe - Jemena Approval # 000413493",
    ))))
    row2 = next(r for r in rows2 if r.legacy_reference == "TESTIMP0009")
    assert row2.parsed["customer_name"] == "Jane Roe"
    assert "jemena" not in row2.parsed["customer_name"].lower()
    n2 = row2.parsed["details"]["notes"]
    cn_review = [m["text"] for m in n2.get("review_notes", []) if m["source_column"] == "Customer Name"]
    cn_misfiled = [m["text"] for m in n2.get("misfiled", []) if m["source_column"] == "Customer Name"]
    assert any("Approval" in t for t in cn_review)
    assert not any("Approval" in t for t in cn_misfiled)

    # An APPROVED status phrase still drives approval_state (unchanged) AND its text
    # rides in neutral review_notes, never misfiled.
    rows4 = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Pat Lee - ERGON APPROVED",
    ))))
    row4 = next(r for r in rows4 if r.legacy_reference == "TESTIMP0009")
    assert row4.parsed["approval_state"] == "approved"
    n4 = row4.parsed["details"]["notes"]
    assert any("APPROVED" in m["text"].upper() for m in n4.get("review_notes", []))
    assert not any(m["source_column"] == "Customer Name" for m in n4.get("misfiled", []))

    # No DOB field is ever invented from a 'dob …' salesperson suffix; it is a
    # neutral review note (not misfiled), and never a sale_date.
    rows3 = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Dana Fox", sales_cell="Robert W - dob 14/05/1980",
    ))))
    row3 = next(r for r in rows3 if r.legacy_reference == "TESTIMP0009")
    assert row3.parsed["salesperson"] == "Robert W"
    assert row3.parsed["sale_date"] is None
    assert "dob" not in str(row3.parsed["details"].get("sales", {})).lower()
    n3 = row3.parsed["details"]["notes"]
    sc_review = [m["text"] for m in n3.get("review_notes", []) if m["source_column"] == "Sales Consultant"]
    assert sc_review == ["dob 14/05/1980"]
    assert not any(m["source_column"] == "Sales Consultant" for m in n3.get("misfiled", []))


# --------------------------------------------------------------------------- #
# Conservative AU address parsing (Phase-7 cleanup). Fabricated addresses only.
# --------------------------------------------------------------------------- #
def test_parse_address_standard_au_format():
    pa = import_parser.parse_address
    r = pa("39 Example St, Cooma NSW 2866")
    assert r["line1"] == "39 Example St"
    assert r["suburb"] == "Cooma"
    assert r["state"] == "NSW"
    assert r["postcode"] == "2866"
    assert r["structured"] is True
    # lowercase state is normalised; reconstruct loses nothing
    r2 = pa("12 Test Road, Sometown vic 3434")
    assert r2["state"] == "VIC" and r2["suburb"] == "Sometown" and r2["postcode"] == "3434"


def test_parse_address_reversed_postcode_state():
    r = import_parser.parse_address("4 Test Place, Bartown 2705 NSW")
    assert r["state"] == "NSW" and r["postcode"] == "2705"
    assert r["suburb"] == "Bartown" and r["line1"] == "4 Test Place"


def test_parse_address_preserves_lot_dp_in_line1():
    pa = import_parser.parse_address
    # Lot in parentheses leading the street — stays in line1, exactly preserved.
    r = pa("(Lot 24) 19 Example Lane, Faketown NSW 2711")
    assert r["line1"] == "(Lot 24) 19 Example Lane"
    assert r["suburb"] == "Faketown" and r["state"] == "NSW" and r["postcode"] == "2711"
    # "Lot 34, 4 St, Town STATE PC": the suburb is the LAST comma segment; the Lot
    # descriptor is preserved within line1.
    r2 = pa("Lot 34, 4 Example St, Faketown NSW 2647")
    assert r2["suburb"] == "Faketown"
    assert r2["line1"] == "Lot 34, 4 Example St"
    assert "Lot 34" in r2["line1"]


def test_parse_address_conservative_when_uncertain():
    pa = import_parser.parse_address
    # No comma -> cannot split suburb without guessing; keep head as line1.
    r = pa("39 Example St Cooma NSW 2866")
    assert r["state"] == "NSW" and r["postcode"] == "2866"
    assert r["line1"] == "39 Example St Cooma" and r["suburb"] is None
    assert r["structured"] is True
    # No state+postcode anchor at all -> keep raw verbatim, structure nothing.
    weird = pa("c/- the shed out the back, ask for directions")
    assert weird["structured"] is False
    assert weird["line1"] == "c/- the shed out the back, ask for directions"
    assert weird["suburb"] is None and weird["state"] is None and weird["postcode"] is None
    # Blank.
    blank = pa("")
    assert blank == {
        "line1": None, "suburb": None, "state": None, "postcode": None,
        "structured": False, "note": None,
    }


def test_parse_rows_attaches_address_parts():
    rows = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(name_cell="Pat Lee"))))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    # The fixture's address is "9 Test St" (no state/pc) -> conservatively unstructured.
    ap = row.parsed["address_parts"]
    assert ap["line1"] == "9 Test St" and ap["structured"] is False
    assert row.parsed["address"] == "9 Test St"  # raw retained for back-compat


# --------------------------------------------------------------------------- #
# F: trailing non-address note peeled off the Address cell
# --------------------------------------------------------------------------- #
def test_parse_address_peels_trailing_billing_note():
    # The exact reported case: the billing note is peeled, the address parses cleanly,
    # and the note is preserved (the raw cell keeps the full original verbatim).
    r = import_parser.parse_address("17 Daalbata Rd , Leeton 2705 NSW - 405 for the bill")
    assert r["line1"] == "17 Daalbata Rd"
    assert r["suburb"] == "Leeton"
    assert r["state"] == "NSW"
    assert r["postcode"] == "2705"
    assert r["structured"] is True
    assert r["note"] == "405 for the bill"          # preserved, NOT part of the address


def test_parse_address_no_trailing_note_unchanged():
    # A normal address with no trailing note: behaviour unchanged, note is None.
    r = import_parser.parse_address("39 Example St, Cooma NSW 2866")
    assert r["line1"] == "39 Example St" and r["suburb"] == "Cooma"
    assert r["state"] == "NSW" and r["postcode"] == "2866" and r["note"] is None


def test_parse_address_does_not_split_hyphen_in_street():
    pa = import_parser.parse_address
    # A hyphen that is part of legitimate street text (before the tail) is NEVER peeled.
    r = pa("5-7 Example St, Town NSW 2000")
    assert r["line1"] == "5-7 Example St" and r["suburb"] == "Town"
    assert r["postcode"] == "2000" and r["note"] is None
    # ...even when a genuine trailing note IS present after the tail: only the note peels.
    r2 = pa("5-7 Example St, Town NSW 2000 - paid in cash")
    assert r2["line1"] == "5-7 Example St" and r2["suburb"] == "Town"
    assert r2["note"] == "paid in cash"


def test_parse_address_preserves_unit_prefix_with_note():
    # A unit prefix stays in line1; a trailing note still peels cleanly.
    r = import_parser.parse_address("Unit 4, 17 Daalbata Rd, Leeton NSW 2705 - 405 for the bill")
    assert "Unit 4" in r["line1"]
    assert r["suburb"] == "Leeton" and r["state"] == "NSW" and r["postcode"] == "2705"
    assert r["note"] == "405 for the bill"


def test_address_note_surfaces_in_imported_notes():
    # The peeled note is preserved as neutral imported review context and surfaces in
    # the seeded internal-notes summary.
    from app.services.import_details import build_details, build_imported_notes

    raw_addr = "17 Daalbata Rd, Leeton 2705 NSW - 405 for the bill"
    parsed = {"address_parts": import_parser.parse_address(raw_addr)}
    details = build_details(parsed, {"address": raw_addr})
    review_texts = [m["text"] for m in (details.get("notes", {}).get("review_notes") or [])]
    assert "405 for the bill" in review_texts
    notes = build_imported_notes(details)
    assert notes is not None and "405 for the bill" in notes


def test_worst_row_combined_pollution_all_preserved():
    """A deliberately convoluted row (mirrors the architecture-pass stress
    categories): polluted Sales Consultant + customer-name legal descriptor +
    in-name distributor approval phrase + decommission marker. Every structured
    field stays clean; every stripped suffix is preserved verbatim with the right
    source_column — the land descriptor as a misfiled source note, the DOB
    remainder and approval phrase as neutral review notes; no DOB is ever invented."""
    rows = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Bill (William) Corn -Lot 4 DP 588479 - JEMENA APPROVAL #000445604 DECOM",
        sales_cell="Robert W - dob 15/06/1965",
    ))))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    p = row.parsed
    # structured fields are clean
    assert p["customer_name"] == "Bill (William) Corn"
    assert p["salesperson"] == "Robert W"
    assert p["sale_date"] is None  # the dob date is NEVER taken as a sale date
    # decommission detected + flagged, not left in the name
    assert p["removes_old_system"] is True
    assert "decom" not in p["customer_name"].lower()
    # every stripped suffix preserved verbatim, tagged by source column, in the
    # right bucket: land descriptor -> misfiled source note; DOB remainder +
    # approval phrase -> neutral review notes.
    notes = p["details"]["notes"]
    mis = {(m["source_column"], m["text"]) for m in notes.get("misfiled", [])}
    review = {(m["source_column"], m["text"]) for m in notes.get("review_notes", [])}
    assert ("Sales Consultant", "dob 15/06/1965") in review
    assert any(c == "Customer Name" and "Lot 4 DP 588479" in t for c, t in mis)
    assert any(c == "Customer Name" and "APPROVAL" in t.upper() for c, t in review)
    # the scary "misfiled" bucket holds neither the DOB nor the approval phrase
    assert not any(c == "Sales Consultant" for c, _t in mis)
    assert not any(c == "Customer Name" and "APPROVAL" in _t.upper() for c, _t in mis)
    # no DOB field/concept anywhere
    assert "dob" not in str(p["details"].get("sales", {})).lower()


def test_nonnumeric_panel_count_is_review_note_and_not_an_error():
    # A battery/inverter-only job (non-numeric panel cell) is valid: the text is a
    # neutral review note, NOT a misfiled warning, and the parser raises NO error
    # issue for it — so it can never block a commit as an unresolved error.
    rows = list(import_parser.parse_rows(_ws_from_bytes(_one_job_row_bytes(
        name_cell="Pat Lee", panels="existing system - battery only",
    ))))
    row = next(r for r in rows if r.legacy_reference == "TESTIMP0009")
    notes = row.parsed["details"]["notes"]
    assert any(
        m["source_column"] == "No of Panels" and "battery only" in m["text"]
        for m in notes.get("review_notes", [])
    )
    assert not any(m["source_column"] == "No of Panels" for m in notes.get("misfiled", []))
    assert "panel_count" not in row.parsed["details"].get("system", {})
    # No error-severity issue is produced by the non-numeric panel cell.
    assert not any(i["severity"] == "error" for i in row.issues)


# --------------------------------------------------------------------------- #
# Ingest tests (staging writes only — NO live Customer/Job writes)
# --------------------------------------------------------------------------- #
def test_ingest_creates_staging_and_no_live_records(db_session: Session, users):
    cust_before = db_session.scalar(select(func.count()).select_from(Customer))
    job_before = db_session.scalar(select(func.count()).select_from(Job))

    batch = import_ingest.ingest_bytes(
        db_session,
        file_bytes=_synthetic_bytes(),
        source_filename="synthetic.xlsx",
        created_by_id=users["admin"].id,
    )
    db_session.flush()

    assert batch.status.value == "parsed"
    assert batch.job_rows == 4
    assert batch.divider_rows == 1
    assert batch.blank_rows == 1
    assert batch.total_rows == 6  # rows after the header
    assert batch.file_sha256 and len(batch.file_sha256) == 64
    assert "/" not in batch.source_filename and "\\" not in batch.source_filename

    n_rows = db_session.scalar(
        select(func.count()).select_from(ImportRow).where(ImportRow.batch_id == batch.id)
    )
    n_issues = db_session.scalar(
        select(func.count()).select_from(ImportIssue).where(ImportIssue.batch_id == batch.id)
    )
    assert n_rows == 6
    assert n_issues >= 4
    # All staged rows are un-committed (no live link).
    committed = db_session.scalars(
        select(ImportRow).where(ImportRow.batch_id == batch.id)
    ).all()
    assert all(r.committed_customer_id is None and r.committed_job_id is None for r in committed)

    # Critical: ingest created ZERO live customers/jobs.
    assert db_session.scalar(select(func.count()).select_from(Customer)) == cust_before
    assert db_session.scalar(select(func.count()).select_from(Job)) == job_before


# --------------------------------------------------------------------------- #
# Endpoint tests (admin-only)
# --------------------------------------------------------------------------- #
def _upload(client, data: bytes, name: str = "synthetic.xlsx"):
    return client.post(
        "/api/v1/imports",
        files={"file": (name, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_admin_can_ingest_and_inspect(client_for, users):
    admin = client_for(users["admin"])
    resp = _upload(admin, _synthetic_bytes())
    assert resp.status_code == 201
    batch = resp.json()
    assert batch["status"] == "parsed"
    assert batch["job_rows"] == 4

    # inspect rows
    rows = admin.get(f"/api/v1/imports/{batch['id']}/rows", params={"row_class": "job"}).json()
    assert rows["total"] == 4
    sample = rows["items"][0]
    assert "raw" in sample and "parsed" in sample  # raw preserved verbatim
    assert admin.get(f"/api/v1/imports/{batch['id']}").status_code == 200
    assert admin.get("/api/v1/imports").json()["total"] >= 1


def test_duplicate_file_conflicts(client_for, users):
    admin = client_for(users["admin"])
    data = _synthetic_bytes()
    assert _upload(admin, data).status_code == 201
    dup = client_for(users["admin"])
    assert _upload(dup, data).status_code == 409  # same sha256


def test_non_admin_cannot_ingest_or_inspect(client_for, users):
    support = client_for(users["support"])
    assert _upload(support, _synthetic_bytes()).status_code == 403
    assert support.get("/api/v1/imports").status_code == 403


def test_non_xlsx_rejected(client_for, users):
    admin = client_for(users["admin"])
    resp = admin.post(
        "/api/v1/imports",
        files={"file": ("notes.txt", b"hello", "text/plain")},
    )
    assert resp.status_code == 400
