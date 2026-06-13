#!/usr/bin/env python
"""Dry-run import parser for the SunCentral jobs workbook (COMPLETED sheet).

READ-ONLY ANALYSIS TOOL. This script does NOT write to the database, does NOT
create customers/jobs, and does NOT modify the workbook. It reads an .xlsx file
supplied on the command line, attempts to parse the COMPLETED sheet into
structured fields, and prints a dry-run report (counts, distributions, issues,
and a few sample parsed rows) so we can design the real staging/import safely.

Usage:
    python backend/scripts/import_dryrun.py "/path/to/workbook.xlsx"
    python backend/scripts/import_dryrun.py "/path/to/workbook.xlsx" --sheet COMPLETED
    python backend/scripts/import_dryrun.py "/path/to/workbook.xlsx" --samples 5
    python backend/scripts/import_dryrun.py "/path/to/workbook.xlsx" --json-output out.json

The workbook path is always supplied as an argument — no real path is hardcoded.
If you save JSON output, put it somewhere git-ignored (e.g. ref/ or backups/);
it can contain customer PII and must not be committed.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency guard
    sys.exit("openpyxl is required: pip install openpyxl (it is in backend/requirements.txt)")


# --------------------------------------------------------------------------- #
# Reference rules (owner-provided)
# --------------------------------------------------------------------------- #
# NMI prefix -> distributor. Longer prefixes are tried before shorter ones.
NMI_PREFIX_RULES: list[tuple[str, str]] = [
    ("4001", "NSW Essential"),
    ("4407", "NSW Essential"),
    ("4204", "NSW Essential"),
    ("4508", "NSW Essential"),
    ("4301", "NSW Endeavour"),
    ("6305", "VIC AusNet"),
    ("6306", "VIC AusNet"),
    ("6407", "VIC United"),
    ("6408", "VIC United"),
    ("6203", "VIC Powercor"),
    ("6204", "VIC Powercor"),
    ("6001", "VIC Jemena"),
    ("410", "NSW Ausgrid"),
    ("304", "QLD Ergon"),
    ("305", "QLD Ergon"),
    ("31", "QLD Energex"),
]
NMI_ALNUM_RULES: list[tuple[str, str]] = [("QB", "QLD Energex")]  # checked on raw text

APPROVAL_APPROVED_RE = re.compile(r"\bAPPROVED\b", re.IGNORECASE)
APPROVAL_PENDING_RE = re.compile(r"\bPENDING\b\s*([0-3]?\d[/\-][0-1]?\d[/\-]\d{2,4})?", re.IGNORECASE)
DATE_RE = re.compile(r"([0-3]?\d[/\-][0-1]?\d[/\-]\d{2,4})")
DOB_RE = re.compile(r"\bDOB\b", re.IGNORECASE)
# Markers after which the Customer Name cell stops being the name.
NAME_STOP_MARKERS = [" - ", " DOB", " PENDING", " APPROVED", " Approved", " LOT", " Lot",
                     " lot", " #", " REF", " Ref", " DP ", " dp "]

DIVIDER_HINTS = ("FORTNIGHT", "WEEK ", "WEEK ", "BELOW", "ABOVE", "MONTH", "TBC")

PANEL_BRAND_HINTS = ("longi", "trina", "ae", "tw", "jinko", "ja ", "canadian", "risen", "qcell",
                     "rec", "sunpower", "hyundai", "seraphim", "phono")
INVERTER_BRAND_HINTS = ("goodwe", "sungrow", "solis", "saj", "alpha", "sigenergy", "solax", "fronius",
                        "growatt", "huawei", "tesla", "sma", "enphase", "redback")


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def s(v: Any) -> str:
    """Normalise a cell value to a trimmed string ('' for None)."""
    if v is None:
        return ""
    text = str(v).strip()
    # Excel stores integers from text columns as floats: '62034764986.0'.
    if re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]
    return text


def infer_distributor(nmi_raw: str) -> tuple[str | None, str | None]:
    """Return (distributor, matched_prefix) inferred from an NMI, or (None, None)."""
    raw = nmi_raw.strip()
    if not raw or raw in {"-", "N/A", "n/a"}:
        return None, None
    upper = raw.upper()
    for prefix, name in NMI_ALNUM_RULES:
        if upper.startswith(prefix):
            return name, prefix
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return None, None
    # Longest specific prefix first.
    for prefix, name in sorted(NMI_PREFIX_RULES, key=lambda x: -len(x[0])):
        if digits.startswith(prefix):
            return name, prefix
    if digits.startswith("2"):  # SA SAPN (single-digit rule, lowest priority)
        return "SA SAPN", "2"
    return None, None


def parse_msb(raw: str) -> str:
    """Map MSB/SB pics cell to yes / maybe / no (real value set is broader than 3)."""
    t = raw.strip().lower()
    if t in {"", "no", "requested"}:
        return "no"
    if t in {"yes?", "??", "?"}:
        return "maybe"
    if t.startswith("yes") or "in drive" in t or "in file" in t or "drive" in t:
        return "yes"
    return "maybe"  # anything unexpected -> maybe, raw preserved by caller


def parse_sales_consultant(raw: str) -> dict[str, Any]:
    """Split 'Example Person 10/10/2025' / 'Sales Rep - 30/06/2023' into name + sale_date."""
    if not raw:
        return {"name": "", "sale_date": None}
    date_m = DATE_RE.search(raw)
    sale_date = date_m.group(1) if date_m else None
    name = raw
    if date_m:
        name = (raw[: date_m.start()] + raw[date_m.end():])
    name = name.strip().rstrip("-").strip().strip("-").strip()
    return {"name": name, "sale_date": sale_date}


def parse_customer_name(raw: str) -> dict[str, Any]:
    """Isolate the customer name; everything after the first marker -> extracted notes."""
    if not raw:
        return {"name": "", "extracted": "", "looks_like_name": False}
    idxs = [raw.find(m) for m in NAME_STOP_MARKERS if raw.find(m) > 0]
    cut = min(idxs) if idxs else len(raw)
    name = raw[:cut].strip()
    extracted = raw[cut:].strip(" -")
    # Heuristic: a real name starts with a letter and isn't a pure reference token.
    looks_like_name = bool(name) and name[0].isalpha() and not re.match(
        r"(?i)^(ref|essential|ergon|energex|approved|pending|lot)\b", name
    )
    return {"name": name, "extracted": extracted, "looks_like_name": looks_like_name}


def parse_approval(*texts: str) -> dict[str, Any]:
    """Detect approval state (approved / pending+date / none) from one or more cells."""
    blob = " ".join(t for t in texts if t)
    if APPROVAL_APPROVED_RE.search(blob):
        return {"state": "approved", "pending_date": None}
    pend = APPROVAL_PENDING_RE.search(blob)
    if pend:
        return {"state": "pending", "pending_date": pend.group(1)}
    return {"state": "none", "pending_date": None}


def parse_phones(raw: str) -> dict[str, Any]:
    """Split phone cell into numbers, keeping a label ONLY when explicit. Never guess owner."""
    if not raw:
        return {"numbers": [], "labelled": False}
    parts = re.split(r"[/]", raw)
    out: list[dict[str, str]] = []
    labelled = False
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # number(s) + optional trailing label after ' - ' or a name
        nums = re.findall(r"\+?\d[\d\s]{6,}\d", part)
        label = re.sub(r"\+?\d[\d\s]{6,}\d", "", part).strip(" -")
        if not nums:
            continue
        explicit = bool(label) and any(ch.isalpha() for ch in label)
        if explicit:
            labelled = True
        for n in nums:
            out.append({"number": re.sub(r"\s+", "", n), "label": label if explicit else ""})
    return {"numbers": out, "labelled": labelled}


def parse_emails(raw: str) -> list[str]:
    if not raw:
        return []
    return [e.strip() for e in raw.split("/") if e.strip() and e.strip().lower() not in {"n/a", "na"}]


def parse_date_maybe(raw: str) -> date | None:
    """Parse an install-date cell. Excel coerces some d/m/y values to ISO datetimes
    (often with a US m/d interpretation), so this is best-effort for the day-vs-date
    contradiction check only; the raw string is always preserved separately."""
    raw = raw.strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", raw)  # Excel-coerced datetime
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"([0-3]?\d)[/\-]([0-1]?\d)[/\-](\d{2,4})", raw)  # dd/mm/yyyy (AU)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y += 2000 if y < 100 else 0
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def hardware_confidence(raw: str, brand_hints: tuple[str, ...]) -> str:
    """Placeholder confidence: confident / uncertain / none."""
    t = raw.strip().lower()
    if t in {"", "-", "/", "0", "n/a"}:
        return "none"
    has_brand = any(h in t for h in brand_hints)
    has_number = bool(re.search(r"\d", t))
    if has_brand and has_number:
        return "confident"
    if has_brand or has_number:
        return "uncertain"
    return "uncertain"


# --------------------------------------------------------------------------- #
# Header detection + column mapping
# --------------------------------------------------------------------------- #
WANTED_HEADERS = {
    "sales consultant": "sales_consultant",
    "customer name": "customer_name",
    "address": "address",
    "phone": "phone",
    "notes": "notes",
    "msb/sb pics in file": "msb",  # header normalisation strips the trailing '?'
    "email": "email",
    "distributor": "distributor",
    "retailer": "retailer",
    "nmi": "nmi",
    "meter no": "meter_no",
    "no of panels": "no_of_panels",
    "panel brand/ wattage": "panel_brand",
    "inverter brand/model": "inverter",
    "storey": "storey",
    "phase": "phase",
    "roof type": "roof_type",
    "date": "date",
    "day": "day",
    "time": "time",
    "installer": "installer",
    "welcome call": "welcome_call",
    "total": "total",
    "deposit": "deposit",
    "balance": "balance",
    "result of payment": "result_of_payment",
    "notes on payment": "notes_on_payment",
    "accreditation code": "accreditation_code",
}


def find_header_row(ws, max_scan: int = 6) -> int | None:
    for r in range(1, max_scan + 1):
        seen = {s(ws.cell(r, c).value).lower().rstrip("?: ").strip() for c in range(1, ws.max_column + 1)}
        if "customer name" in seen and "nmi" in seen:
            return r
    return None


def build_colmap(ws, header_row: int) -> dict[str, int]:
    colmap: dict[str, int] = {"ref": 1}  # first column = legacy reference
    for c in range(1, ws.max_column + 1):
        key = s(ws.cell(header_row, c).value).lower().rstrip("?: ").strip()
        if key in WANTED_HEADERS:
            colmap[WANTED_HEADERS[key]] = c
    return colmap


# --------------------------------------------------------------------------- #
# Row classification + parsing
# --------------------------------------------------------------------------- #
REF_RE = re.compile(r"^SCS?\d{3,4}\b", re.IGNORECASE)


@dataclass
class Report:
    sheet: str = ""
    classes: Counter = field(default_factory=Counter)
    sales: Counter = field(default_factory=Counter)
    installers: Counter = field(default_factory=Counter)
    distributors_raw: Counter = field(default_factory=Counter)
    retailers_raw: Counter = field(default_factory=Counter)
    nmi_inferred: Counter = field(default_factory=Counter)
    approval: Counter = field(default_factory=Counter)
    msb: Counter = field(default_factory=Counter)
    issues: Counter = field(default_factory=Counter)
    issue_samples: list[dict] = field(default_factory=list)
    parsed_samples: list[dict] = field(default_factory=list)
    likely_jobs: int = 0


def classify(ref: str, nonempty: int, parsed_name: dict) -> str:
    if nonempty == 0:
        return "blank"
    if ref and not REF_RE.match(ref):
        if any(h in ref.upper() for h in DIVIDER_HINTS) or nonempty <= 2:
            return "divider"
    if REF_RE.match(ref):
        return "job"
    if parsed_name["name"] and nonempty <= 3:
        return "ambiguous"
    if nonempty >= 5:
        return "job"  # data-bearing row without a clean ref
    return "ambiguous"


def run(path: str, sheet: str, samples: int) -> tuple[Report, list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    header_row = find_header_row(ws)
    if header_row is None:
        sys.exit("Could not locate a header row (expected 'Customer Name' + 'NMI').")
    cm = build_colmap(ws, header_row)

    def get(row: int, key: str) -> str:
        c = cm.get(key)
        return s(ws.cell(row, c).value) if c else ""

    rep = Report(sheet=sheet)
    all_parsed: list[dict] = []
    current_context = ""

    for r in range(header_row + 1, ws.max_row + 1):
        nonempty = sum(1 for c in range(1, min(ws.max_column, 40) + 1) if s(ws.cell(r, c).value))
        ref = get(r, "ref")
        name_info = parse_customer_name(get(r, "customer_name"))
        klass = classify(ref, nonempty, name_info)
        rep.classes[klass] += 1

        if klass == "blank":
            continue
        if klass == "divider":
            current_context = ref or name_info["name"]
            continue

        # --- parse a (likely) job / ambiguous data row ---
        sales = parse_sales_consultant(get(r, "sales_consultant"))
        phones = parse_phones(get(r, "phone"))
        emails = parse_emails(get(r, "email"))
        nmi_raw = get(r, "nmi")
        dist_inferred, prefix = infer_distributor(nmi_raw)
        dist_raw = get(r, "distributor")
        retailer_raw = get(r, "retailer")
        approval = parse_approval(name_info["extracted"], get(r, "notes"))
        msb_raw = get(r, "msb")
        msb_state = parse_msb(msb_raw)
        panel_raw = get(r, "panel_brand")
        inverter_raw = get(r, "inverter")

        parsed = {
            "row": r,
            "class": klass,
            "legacy_reference": ref,  # exact text preserved
            "context_from_divider": current_context or None,
            "salesperson": sales["name"],
            "sale_date": sales["sale_date"],
            "customer_name": name_info["name"],
            "name_extracted_notes": name_info["extracted"] or None,
            "approval_state": approval["state"],
            "approval_pending_date": approval["pending_date"],
            "phones": phones["numbers"],
            "emails": emails,
            "msb_state": msb_state,
            "msb_raw": msb_raw or None,
            "distributor_raw": dist_raw or None,
            "distributor_inferred": dist_inferred,
            "retailer_raw": retailer_raw or None,
            "nmi_raw": nmi_raw or None,
            "meter_no": get(r, "meter_no") or None,
            "no_of_panels": get(r, "no_of_panels") or None,
            "panel_raw": panel_raw or None,
            "panel_confidence": hardware_confidence(panel_raw, PANEL_BRAND_HINTS),
            "inverter_raw": inverter_raw or None,
            "inverter_confidence": hardware_confidence(inverter_raw, INVERTER_BRAND_HINTS),
            "install_date": get(r, "date") or None,
            "install_day": get(r, "day") or None,
            "install_time": get(r, "time") or None,
            "installer_raw": get(r, "installer") or None,
            "payment": {
                "total": get(r, "total") or None,
                "deposit": get(r, "deposit") or None,
                "balance": get(r, "balance") or None,
                "result": get(r, "result_of_payment") or None,
                "notes": get(r, "notes_on_payment") or None,
            },
            "compliance": {
                "accreditation_code": get(r, "accreditation_code") or None,
                "welcome_call": get(r, "welcome_call") or None,
            },
            "notes_raw": get(r, "notes") or None,
        }
        all_parsed.append(parsed)
        if klass == "job":
            rep.likely_jobs += 1

        # --- aggregate distributions ---
        if sales["name"]:
            rep.sales[sales["name"]] += 1
        if parsed["installer_raw"]:
            rep.installers[parsed["installer_raw"]] += 1
        if dist_raw:
            rep.distributors_raw[dist_raw] += 1
        if retailer_raw:
            rep.retailers_raw[retailer_raw] += 1
        rep.nmi_inferred[dist_inferred or "UNMATCHED/none"] += 1
        rep.approval[approval["state"]] += 1
        rep.msb[msb_state] += 1

        # --- issues ---
        def add_issue(kind: str, reason: str) -> None:
            rep.issues[kind] += 1
            if len(rep.issue_samples) < 40:
                rep.issue_samples.append({"row": r, "kind": kind, "reason": reason})

        if not name_info["looks_like_name"]:
            add_issue("ambiguous_name", f"name={name_info['name']!r}")
        if len(phones["numbers"]) > 1:
            add_issue("multi_phone", f"{len(phones['numbers'])} numbers")
        if len(emails) > 1:
            add_issue("multi_email", f"{len(emails)} emails")
        if nmi_raw and dist_inferred is None:
            add_issue("nmi_unmatched", f"nmi={nmi_raw!r}")
        if dist_raw and dist_inferred and dist_inferred.split()[-1].lower() not in dist_raw.lower():
            add_issue("distributor_mismatch", f"sheet={dist_raw!r} vs inferred={dist_inferred!r}")
        if parsed["inverter_confidence"] == "uncertain" and inverter_raw:
            add_issue("hardware_uncertain", f"inverter={inverter_raw[:40]!r}")
        if approval["state"] == "pending" and not approval["pending_date"]:
            add_issue("approval_pending_no_date", "")
        # Day-of-week must not contradict the install date (Excel US-coercion catches here).
        d_parsed = parse_date_maybe(parsed["install_date"] or "")
        if d_parsed and parsed["install_day"]:
            if d_parsed.strftime("%A").lower() != parsed["install_day"].strip().lower():
                add_issue("date_day_mismatch", f"{parsed['install_date']} is {d_parsed.strftime('%A')}, day says {parsed['install_day']!r}")

    rep.parsed_samples = [p for p in all_parsed if p["class"] == "job"][:samples]
    return rep, all_parsed


# --------------------------------------------------------------------------- #
# Report printing
# --------------------------------------------------------------------------- #
def print_report(rep: Report) -> None:
    def section(title: str) -> None:
        print(f"\n=== {title} ===")

    print("DRY-RUN IMPORT REPORT (read-only — no database writes)")
    print(f"Sheet: {rep.sheet!r}")

    section("Row classification")
    for k in ("blank", "divider", "job", "ambiguous"):
        print(f"  {k:10}: {rep.classes.get(k, 0)}")
    print(f"  likely jobs (clean ref): {rep.likely_jobs}")

    section("Approval status")
    for k, v in rep.approval.most_common():
        print(f"  {k:10}: {v}")

    section("MSB/SB pics state")
    for k, v in rep.msb.most_common():
        print(f"  {k:10}: {v}")

    section("NMI -> distributor inference")
    for k, v in rep.nmi_inferred.most_common():
        print(f"  {v:5}  {k}")

    section("Top salesperson (parsed name)")
    for k, v in rep.sales.most_common(10):
        print(f"  {v:4}  {k!r}")

    section("Top installer (raw)")
    for k, v in rep.installers.most_common(12):
        print(f"  {v:4}  {k!r}")

    section("Top distributor (raw)")
    for k, v in rep.distributors_raw.most_common(8):
        print(f"  {v:4}  {k!r}")

    section("Top retailer (raw)")
    for k, v in rep.retailers_raw.most_common(8):
        print(f"  {v:4}  {k!r}")

    section("Issue counts")
    for k, v in rep.issues.most_common():
        print(f"  {v:5}  {k}")

    section("Sample issues (row · kind · reason)")
    for it in rep.issue_samples[:15]:
        print(f"  row {it['row']}: {it['kind']} — {it['reason']}")

    section("Sample parsed job rows")
    for p in rep.parsed_samples:
        print(f"\n  row {p['row']} ref={p['legacy_reference']!r}")
        print(f"    customer: {p['customer_name']!r}  extracted_notes: {p['name_extracted_notes']!r}")
        print(f"    sales: {p['salesperson']!r} sale_date={p['sale_date']}  install={p['install_date']} ({p['install_day']})")
        print(f"    phones: {p['phones']}")
        print(f"    emails: {p['emails']}")
        print(f"    approval: {p['approval_state']} {p['approval_pending_date'] or ''}  msb: {p['msb_state']} (raw {p['msb_raw']!r})")
        print(f"    nmi: {p['nmi_raw']!r} -> inferred {p['distributor_inferred']!r}; sheet dist {p['distributor_raw']!r}; retailer {p['retailer_raw'] if 'retailer_raw' in p else ''}")
        print(f"    panels: {p['no_of_panels']} {p['panel_raw']!r} [{p['panel_confidence']}]")
        print(f"    inverter: {p['inverter_raw']!r} [{p['inverter_confidence']}]")
        print(f"    installer: {p['installer_raw']!r}  context: {p['context_from_divider']!r}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Read-only dry-run parser for the jobs workbook.")
    ap.add_argument("workbook", help="Path to the .xlsx workbook (never hardcoded).")
    ap.add_argument("--sheet", default="COMPLETED", help="Sheet to analyse (default: COMPLETED).")
    ap.add_argument("--samples", type=int, default=5, help="Number of sample parsed rows to show.")
    ap.add_argument("--json-output", default=None, help="Optional path to write full parsed JSON (PII — keep git-ignored).")
    args = ap.parse_args()

    rep, all_parsed = run(args.workbook, args.sheet, args.samples)
    print_report(rep)

    if args.json_output:
        with open(args.json_output, "w", encoding="utf-8") as fh:
            json.dump({"sheet": rep.sheet, "rows": all_parsed}, fh, indent=2, ensure_ascii=False)
        print(f"\n[wrote full parsed JSON -> {args.json_output} — contains PII, keep it git-ignored]")


if __name__ == "__main__":
    main()
