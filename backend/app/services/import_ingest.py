"""Parse-only ingest of a workbook sheet into staging tables (Phase A).

Creates ImportBatch / ImportRow / ImportIssue rows ONLY. This service never
touches Customer, Job, Task, Activity, Document, or NAS. Committing reviewed rows
to live tables is a separate, future, separately-approved phase.
"""

from __future__ import annotations

import hashlib
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.enums import ImportBatchStatus, ImportRowClass
from app.models.import_staging import ImportBatch, ImportIssue, ImportRow
from app.services import import_parser
from app.services.import_hardware import enrich_row_hardware

DEFAULT_SHEET = "COMPLETED"


def sha256_of(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def load_sheet(file_bytes: bytes, sheet_name: str):
    """Load a worksheet from in-memory xlsx bytes (read-only, data-only)."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    return wb[sheet_name]


def find_duplicate_batch(db: Session, file_sha256: str) -> ImportBatch | None:
    """Return a prior non-deleted batch with the same file hash, if any."""
    return db.scalar(
        select(ImportBatch).where(
            ImportBatch.file_sha256 == file_sha256, ImportBatch.deleted_at.is_(None)
        )
    )


def ingest_worksheet(
    db: Session,
    *,
    ws,
    source_filename: str,
    sheet_name: str,
    file_sha256: str | None,
    created_by_id: int | None,
) -> ImportBatch:
    """Parse a worksheet into staging rows + issues. Adds (does not commit).

    Returns the ImportBatch with `status=parsed`. No live records are created.
    """
    batch = ImportBatch(
        source_filename=source_filename,
        sheet_name=sheet_name,
        file_sha256=file_sha256,
        status=ImportBatchStatus.PARSING,
        created_by_id=created_by_id,
    )
    db.add(batch)
    db.flush()  # assign batch.id

    counts = {ImportRowClass.BLANK: 0, ImportRowClass.DIVIDER: 0,
              ImportRowClass.JOB: 0, ImportRowClass.AMBIGUOUS: 0}
    total = 0
    issue_total = 0

    for prow in import_parser.parse_rows(ws):
        total += 1
        counts[ImportRowClass(prow.row_class)] += 1
        # Stage 4B: DB-aware hardware enrichment (the pure parser stays DB-free). Parses the row's
        # hardware cells ONCE into parsed['details']['hardware'] so preview/review + commit read the
        # SAME stored snapshot. Read-only against the catalogue; legacy details.system.* untouched.
        enrich_row_hardware(db, prow.parsed)
        row = ImportRow(
            batch_id=batch.id,
            source_row_index=prow.source_row_index,
            row_class=ImportRowClass(prow.row_class),
            legacy_reference=prow.legacy_reference or None,
            raw=prow.raw,
            parsed=prow.parsed or None,
            context_text=prow.context_text,
        )
        db.add(row)
        db.flush()  # assign row.id for issue FK
        for issue in prow.issues:
            issue_total += 1
            db.add(
                ImportIssue(
                    row_id=row.id,
                    batch_id=batch.id,
                    kind=issue["kind"],
                    severity=issue["severity"],
                    field=issue.get("field"),
                    message=issue["message"],
                )
            )

    batch.total_rows = total
    batch.job_rows = counts[ImportRowClass.JOB]
    batch.divider_rows = counts[ImportRowClass.DIVIDER]
    batch.blank_rows = counts[ImportRowClass.BLANK]
    batch.ambiguous_rows = counts[ImportRowClass.AMBIGUOUS]
    batch.issue_count = issue_total
    batch.status = ImportBatchStatus.PARSED
    return batch


def ingest_bytes(
    db: Session,
    *,
    file_bytes: bytes,
    source_filename: str,
    sheet_name: str = DEFAULT_SHEET,
    created_by_id: int | None,
) -> ImportBatch:
    """Convenience: hash + load + ingest from raw xlsx bytes (adds, no commit)."""
    digest = sha256_of(file_bytes)
    ws = load_sheet(file_bytes, sheet_name)
    return ingest_worksheet(
        db,
        ws=ws,
        source_filename=source_filename,
        sheet_name=sheet_name,
        file_sha256=digest,
        created_by_id=created_by_id,
    )
